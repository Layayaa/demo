"""
生成Excel模板文件（使用openpyxl，无需pandas）
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
import os

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "询价记录"

# 定义表头
headers = [
    '序号', '项目名称', '材料名称', '规格型号', '单位', '单价',
    '是否含税', '供应商/来源', '地区', '报价时间', '备注',
    '填报部门', '填报工程师', '上传人', '询价类别'
]

# 写入表头
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 示例数据
sample_data = [
    [1, '成都天府新区道路改造项目', '螺纹钢', 'HRB400E Φ16', '吨', 4250, '是', '四川攀钢集团', '成都',
     (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'), '项目急需', '采购部', '张三', '张三', '项目询价'],
    [2, '成都天府新区道路改造项目', '水泥', 'P.O 42.5', '吨', 480, '是', '峨眉山水泥厂', '成都',
     (datetime.now() - timedelta(days=25)).strftime('%Y-%m-%d'), '批量采购', '采购部', '张三', '张三', '项目询价'],
    [3, '绵阳科技城建设项目', '砂石', '中砂', '方', 85, '是', '本地砂石场', '绵阳',
     (datetime.now() - timedelta(days=60)).strftime('%Y-%m-%d'), '', '工程部', '李四', '李四', '项目询价'],
    [4, '绵阳科技城建设项目', '混凝土', 'C30', '方', 420, '是', '成都混凝土公司', '绵阳',
     (datetime.now() - timedelta(days=55)).strftime('%Y-%m-%d'), '现货供应', '工程部', '李四', '李四', '项目询价'],
    [5, '德阳工业园区基础设施项目', '沥青', 'AH-70', '吨', 3850, '是', '四川沥青厂', '德阳',
     (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d'), '', '采购部', '王五', '王五', '项目询价'],
    [6, '德阳工业园区基础设施项目', '管材', 'DN300 钢管', '米', 280, '是', '成都钢管厂', '德阳',
     (datetime.now() - timedelta(days=85)).strftime('%Y-%m-%d'), '定制规格', '工程部', '王五', '王五', '项目询价'],
    [7, '宜宾临港经济开发区项目', '电缆', 'YJV-4×185', '米', 185, '是', '四川电缆厂', '宜宾',
     (datetime.now() - timedelta(days=120)).strftime('%Y-%m-%d'), '', '采购部', '赵六', '赵六', '项目询价'],
    [8, '宜宾临港经济开发区项目', '灯具', 'LED路灯 150W', '套', 850, '是', '深圳照明公司', '宜宾',
     (datetime.now() - timedelta(days=115)).strftime('%Y-%m-%d'), '含安装', '工程部', '赵六', '赵六', '项目询价'],
    [9, '泸州长江大桥改造项目', '涂料', '外墙乳胶漆', '桶', 380, '是', '成都涂料厂', '泸州',
     (datetime.now() - timedelta(days=150)).strftime('%Y-%m-%d'), '', '采购部', '钱七', '钱七', '项目询价'],
    [10, '泸州长江大桥改造项目', '保温材料', '岩棉板 50mm', '平米', 45, '是', '四川保温材料厂', '泸州',
     (datetime.now() - timedelta(days=145)).strftime('%Y-%m-%d'), '防火等级A级', '工程部', '钱七', '钱七', '项目询价'],
]

# 写入示例数据
for row_idx, row_data in enumerate(sample_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center")

# 调整列宽
column_widths = {
    'A': 6, 'B': 25, 'C': 15, 'D': 18, 'E': 8, 'F': 10,
    'G': 10, 'H': 18, 'I': 10, 'J': 12, 'K': 15,
    'L': 12, 'M': 12, 'N': 10, 'O': 12
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# 保存文件
output_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'sample_data')
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

output_path = os.path.join(output_dir, '询价模板_示例数据.xlsx')
wb.save(output_path)

print(f"模板文件已生成: {output_path}")
print(f"包含 10 条示例数据")
print("\n字段列表:")
for header in headers:
    print(f"  - {header}")
